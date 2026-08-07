
from flask import Blueprint, request, jsonify, send_file, current_app
from flask_jwt_extended import jwt_required, get_jwt_identity
from extensions import db, cache
from models import User, Event, Category, Booking, Payment, ActivityLog, Notification
from datetime import datetime, timedelta
from sqlalchemy import func, and_, or_
import json
import csv
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from utils.helpers import admin_required

admin_bp = Blueprint('admin', __name__)

# ==================== DASHBOARD ====================

@admin_bp.route('/dashboard', methods=['GET'])
@jwt_required()
@admin_required
@cache.cached(timeout=300)
def get_dashboard_stats():
    # Basic stats
    total_users = User.query.count()
    total_events = Event.query.count()
    active_events = Event.query.filter_by(status='approved').count()
    completed_events = Event.query.filter_by(status='completed').count()
    cancelled_events = Event.query.filter_by(status='cancelled').count()
    total_bookings = Booking.query.filter_by(booking_status='confirmed').count()
    
    # Revenue
    revenue = db.session.query(func.sum(Booking.total_amount)).filter(
        Booking.booking_status == 'confirmed'
    ).scalar() or 0
    
    # Total payments
    total_payments = db.session.query(func.sum(Payment.amount)).filter(
        Payment.payment_status == 'completed'
    ).scalar() or 0
    
    # Recent users (last 10)
    recent_users = User.query.order_by(User.created_at.desc()).limit(10).all()
    
    # Recent bookings (last 10)
    recent_bookings = Booking.query.order_by(Booking.booking_date.desc()).limit(10).all()
    
    # Popular events (top 5)
    popular_events = db.session.query(
        Event.id,
        Event.title,
        func.count(Booking.id).label('booking_count')
    ).outerjoin(Booking).filter(
        Booking.booking_status == 'confirmed'
    ).group_by(Event.id).order_by(
        func.count(Booking.id).desc()
    ).limit(5).all()
    
    # Monthly bookings (last 6 months)
    monthly_bookings = db.session.query(
        func.strftime('%Y-%m', Booking.booking_date).label('month'),
        func.count(Booking.id).label('count')
    ).filter(
        Booking.booking_date >= datetime.utcnow() - timedelta(days=180),
        Booking.booking_status == 'confirmed'
    ).group_by(func.strftime('%Y-%m', Booking.booking_date)).order_by('month').all()
    
    # Monthly revenue (last 6 months)
    monthly_revenue = db.session.query(
        func.strftime('%Y-%m', Booking.booking_date).label('month'),
        func.sum(Booking.total_amount).label('revenue')
    ).filter(
        Booking.booking_date >= datetime.utcnow() - timedelta(days=180),
        Booking.booking_status == 'confirmed'
    ).group_by(func.strftime('%Y-%m', Booking.booking_date)).order_by('month').all()
    
    # Category stats
    category_stats = db.session.query(
        Category.name,
        func.count(Event.id).label('count')
    ).outerjoin(Event).group_by(Category.id).all()
    
    # User growth (last 6 months)
    user_growth = db.session.query(
        func.strftime('%Y-%m', User.created_at).label('month'),
        func.count(User.id).label('count')
    ).filter(
        User.created_at >= datetime.utcnow() - timedelta(days=180)
    ).group_by(func.strftime('%Y-%m', User.created_at)).order_by('month').all()
    
    return jsonify({
        'total_users': total_users,
        'total_events': total_events,
        'active_events': active_events,
        'completed_events': completed_events,
        'cancelled_events': cancelled_events,
        'total_bookings': total_bookings,
        'revenue': revenue,
        'total_payments': total_payments,
        'recent_users': [user.to_dict() for user in recent_users],
        'recent_bookings': [booking.to_dict() for booking in recent_bookings],
        'popular_events': [
            {'id': e.id, 'title': e.title, 'booking_count': e.booking_count}
            for e in popular_events
        ],
        'monthly_bookings': [
            {'month': m.month, 'count': m.count}
            for m in monthly_bookings
        ],
        'monthly_revenue': [
            {'month': m.month, 'revenue': m.revenue}
            for m in monthly_revenue
        ],
        'category_stats': [
            {'name': c.name, 'count': c.count}
            for c in category_stats
        ],
        'user_growth': [
            {'month': u.month, 'count': u.count}
            for u in user_growth
        ]
    }), 200

# ==================== USERS ====================

@admin_bp.route('/users', methods=['GET'])
@jwt_required()
@admin_required
def get_users():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    search = request.args.get('search', '')
    role = request.args.get('role', '')
    status = request.args.get('status', '')
    sort_by = request.args.get('sort_by', 'created_at')
    sort_order = request.args.get('sort_order', 'desc')
    
    query = User.query
    
    if search:
        query = query.filter(
            or_(
                User.username.ilike(f'%{search}%'),
                User.email.ilike(f'%{search}%')
            )
        )
    
    if role:
        query = query.filter_by(role=role)
    
    if status:
        query = query.filter_by(status=status)
    
    # Sorting
    if sort_by == 'created_at':
        order_col = User.created_at
    elif sort_by == 'username':
        order_col = User.username
    elif sort_by == 'email':
        order_col = User.email
    else:
        order_col = User.id
    
    if sort_order == 'desc':
        query = query.order_by(order_col.desc())
    else:
        query = query.order_by(order_col.asc())
    
    paginated = query.paginate(page=page, per_page=per_page, error_out=False)
    
    return jsonify({
        'users': [user.to_dict() for user in paginated.items],
        'total': paginated.total,
        'page': page,
        'pages': paginated.pages,
        'per_page': per_page
    }), 200

@admin_bp.route('/users', methods=['POST'])
@jwt_required()
@admin_required
def create_user():
    """Create a new user (Admin only)"""
    try:
        data = request.get_json()
        
        # Validate required fields
        required_fields = ['username', 'email', 'password']
        for field in required_fields:
            if not data.get(field):
                return jsonify({'error': f'{field} is required'}), 400
        
        # Check if email exists
        if User.query.filter_by(email=data['email']).first():
            return jsonify({'error': 'Email already exists'}), 400
        
        # Check if username exists
        if User.query.filter_by(username=data['username']).first():
            return jsonify({'error': 'Username already exists'}), 400
        
        # Create user
        user = User(
            username=data['username'],
            email=data['email'],
            phone=data.get('phone', ''),
            role=data.get('role', 'user'),
            status=data.get('status', 'active'),
            email_verified=True
        )
        user.set_password(data['password'])
        
        db.session.add(user)
        db.session.commit()
        
        # Log activity
        log = ActivityLog(
            user_id=get_jwt_identity(),
            action='create_user',
            resource_type='user',
            resource_id=user.id,
            details=f"Admin created user {user.username}",
            ip_address=request.remote_addr,
            user_agent=request.headers.get('User-Agent')
        )
        db.session.add(log)
        db.session.commit()
        
        return jsonify({
            'message': 'User created successfully',
            'user': user.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': str(e)}), 500

@admin_bp.route('/users/<int:user_id>', methods=['PUT'])
@jwt_required()
@admin_required
def update_user(user_id):
    user = User.query.get(user_id)
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    if user.id == get_jwt_identity():
        return jsonify({'error': 'Cannot modify your own account via this endpoint'}), 400
    
    data = request.get_json()
    
    if data.get('status'):
        user.status = data['status']
    
    if data.get('role'):
        user.role = data['role']
    
    db.session.commit()
    cache.clear()
    
    # Send notification to user
    notification = Notification(
        user_id=user.id,
        title="Account Updated",
        message=f"Your account has been updated by admin. Status: {user.status}, Role: {user.role}",
        type='account_update'
    )
    db.session.add(notification)
    db.session.commit()
    
    return jsonify({
        'message': 'User updated successfully',
        'user': user.to_dict()
    }), 200

@admin_bp.route('/users/<int:user_id>', methods=['DELETE'])
@jwt_required()
@admin_required
def delete_user(user_id):
    user = User.query.get(user_id)
    
    if not user:
        return jsonify({'error': 'User not found'}), 404
    
    if user.id == get_jwt_identity():
        return jsonify({'error': 'Cannot delete your own account'}), 400
    
    if user.role == 'admin':
        admin_count = User.query.filter_by(role='admin').count()
        if admin_count <= 1:
            return jsonify({'error': 'Cannot delete the last admin'}), 400
    
    db.session.delete(user)
    db.session.commit()
    cache.clear()
    
    return jsonify({'message': 'User deleted successfully'}), 200

# ==================== EVENTS ====================

@admin_bp.route('/events', methods=['GET'])
@jwt_required()
@admin_required
def get_admin_events():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status', '')
    search = request.args.get('search', '')
    
    query = Event.query
    
    if status:
        query = query.filter_by(status=status)
    
    if search:
        query = query.filter(
            or_(
                Event.title.ilike(f'%{search}%'),
                Event.description.ilike(f'%{search}%')
            )
        )
    
    paginated = query.order_by(Event.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'events': [event.to_dict() for event in paginated.items],
        'total': paginated.total,
        'page': page,
        'pages': paginated.pages,
        'per_page': per_page
    }), 200

@admin_bp.route('/events/<int:event_id>/status', methods=['PUT'])
@jwt_required()
@admin_required
def update_event_status(event_id):
    event = Event.query.get(event_id)
    
    if not event:
        return jsonify({'error': 'Event not found'}), 404
    
    data = request.get_json()
    status = data.get('status')
    
    if not status:
        return jsonify({'error': 'Status is required'}), 400
    
    if status not in ['draft', 'pending', 'approved', 'rejected', 'cancelled', 'completed']:
        return jsonify({'error': 'Invalid status'}), 400
    
    old_status = event.status
    event.status = status
    db.session.commit()
    cache.clear()
    
    # Send notification to organizers
    for organizer in event.organizers:
        notification = Notification(
            user_id=organizer.id,
            title=f"Event {status.capitalize()}",
            message=f"Your event '{event.title}' has been {status} by admin",
            type='event_status_update',
            link=f'/events/{event.id}',
            icon='📌'
        )
        db.session.add(notification)
    
    db.session.commit()
    
    # Log activity
    log = ActivityLog(
        user_id=get_jwt_identity(),
        action='update_event_status',
        resource_type='event',
        resource_id=event.id,
        details=f"Status changed from {old_status} to {status}",
        ip_address=request.remote_addr,
        user_agent=request.headers.get('User-Agent')
    )
    db.session.add(log)
    db.session.commit()
    
    return jsonify({
        'message': 'Event status updated successfully',
        'event': event.to_dict()
    }), 200

# ==================== BOOKINGS ====================

@admin_bp.route('/bookings', methods=['GET'])
@jwt_required()
@admin_required
def get_admin_bookings():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    status = request.args.get('status', '')
    event_id = request.args.get('event_id', type=int)
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    
    query = Booking.query
    
    if status:
        query = query.filter_by(booking_status=status)
    
    if event_id:
        query = query.filter_by(event_id=event_id)
    
    if date_from:
        query = query.filter(Booking.booking_date >= datetime.fromisoformat(date_from))
    
    if date_to:
        query = query.filter(Booking.booking_date <= datetime.fromisoformat(date_to))
    
    paginated = query.order_by(Booking.booking_date.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'bookings': [booking.to_dict() for booking in paginated.items],
        'total': paginated.total,
        'page': page,
        'pages': paginated.pages
    }), 200

@admin_bp.route('/bookings/<int:booking_id>', methods=['PUT'])
@jwt_required()
@admin_required
def update_booking(booking_id):
    booking = Booking.query.get(booking_id)
    
    if not booking:
        return jsonify({'error': 'Booking not found'}), 404
    
    data = request.get_json()
    
    if data.get('booking_status'):
        booking.booking_status = data['booking_status']
    
    if data.get('payment_status'):
        booking.payment_status = data['payment_status']
    
    db.session.commit()
    
    # Notify user
    notification = Notification(
        user_id=booking.user_id,
        title="Booking Updated",
        message=f"Your booking for '{booking.event.title}' has been updated to {booking.booking_status}",
        type='booking_update',
        link=f'/bookings/{booking.id}'
    )
    db.session.add(notification)
    db.session.commit()
    
    return jsonify({
        'message': 'Booking updated successfully',
        'booking': booking.to_dict()
    }), 200

# ==================== EXPORT ====================

@admin_bp.route('/export/<string:type>', methods=['GET'])
@jwt_required()
@admin_required
def export_data(type):
    if type not in ['csv', 'excel']:
        return jsonify({'error': 'Invalid export type'}), 400
    
    # Get data
    bookings = Booking.query.filter_by(booking_status='confirmed').order_by(
        Booking.booking_date.desc()
    ).all()
    
    if type == 'csv':
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write headers
        headers = ['Booking ID', 'Event', 'User', 'Quantity', 'Total', 'Date', 'Status']
        writer.writerow(headers)
        
        # Write data
        for booking in bookings:
            writer.writerow([
                booking.id,
                booking.event.title,
                booking.user.username,
                booking.quantity,
                booking.total_amount,
                booking.booking_date.strftime('%Y-%m-%d %H:%M'),
                booking.booking_status
            ])
        
        output.seek(0)
        return send_file(
            io.BytesIO(output.getvalue().encode()),
            as_attachment=True,
            download_name=f'bookings_export_{datetime.utcnow().strftime("%Y%m%d")}.csv',
            mimetype='text/csv'
        )
    
    elif type == 'excel':
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Bookings"
        
        # Style headers
        headers = ['Booking ID', 'Event', 'User', 'Quantity', 'Total', 'Date', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.font = Font(color="FFFFFF")
            cell.alignment = Alignment(horizontal='center')
        
        # Write data
        for row, booking in enumerate(bookings, 2):
            ws.cell(row=row, column=1, value=booking.id)
            ws.cell(row=row, column=2, value=booking.event.title)
            ws.cell(row=row, column=3, value=booking.user.username)
            ws.cell(row=row, column=4, value=booking.quantity)
            ws.cell(row=row, column=5, value=booking.total_amount)
            ws.cell(row=row, column=6, value=booking.booking_date.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row, column=7, value=booking.booking_status)
        
        # Adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        return send_file(
            output,
            as_attachment=True,
            download_name=f'bookings_export_{datetime.utcnow().strftime("%Y%m%d")}.xlsx',
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )

# ==================== ACTIVITY LOGS ====================

@admin_bp.route('/activity-logs', methods=['GET'])
@jwt_required()
@admin_required
def get_activity_logs():
    page = request.args.get('page', 1, type=int)
    per_page = request.args.get('per_page', 20, type=int)
    action = request.args.get('action', '')
    user_id = request.args.get('user_id', type=int)
    
    query = ActivityLog.query
    
    if action:
        query = query.filter_by(action=action)
    
    if user_id:
        query = query.filter_by(user_id=user_id)
    
    paginated = query.order_by(ActivityLog.timestamp.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )
    
    return jsonify({
        'logs': [log.to_dict() for log in paginated.items],
        'total': paginated.total,
        'page': page,
        'pages': paginated.pages
    }), 200